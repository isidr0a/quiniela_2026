import json
from pathlib import Path
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "JESUS QUINIELA.xlsx"
OUTPUT_PATH = ROOT / "data" / "quiniela.json"


def clean_name(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def int_or_none(value):
    if value is None or value == "":
        return None
    return int(value)


def main():
    workbook = load_workbook(XLSX_PATH, data_only=True)
    sheet = workbook.active

    matches = []
    for column in range(2, sheet.max_column + 1, 3):
        match_number = sheet.cell(row=1, column=column).value
        if match_number is None:
            continue

        matches.append(
            {
                "id": int(match_number),
                "label": f"Partido {int(match_number)}",
                "homeTeam": None,
                "awayTeam": None,
                "group": None,
                "kickoff": None,
                "result": {
                    "home": None,
                    "away": None,
                    "status": "pending"
                },
                "sourceColumn": {
                    "homePrediction": column,
                    "awayPrediction": column + 1,
                    "points": column + 2
                }
            }
        )

    participants = []
    for row in range(3, sheet.max_row + 1):
        name = clean_name(sheet.cell(row=row, column=1).value)
        if not name:
            continue

        predictions = []
        for match in matches:
            column = match["sourceColumn"]["homePrediction"]
            home = int_or_none(sheet.cell(row=row, column=column).value)
            away = int_or_none(sheet.cell(row=row, column=column + 1).value)
            points = int_or_none(sheet.cell(row=row, column=column + 2).value)

            predictions.append(
                {
                    "matchId": match["id"],
                    "home": home,
                    "away": away,
                    "points": points
                }
            )

        participants.append(
            {
                "id": name.lower().replace(" ", "-"),
                "name": name,
                "predictions": predictions
            }
        )

    payload = {
        "source": {
            "file": XLSX_PATH.name,
            "sheet": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column
        },
        "scoring": {
            "exactScoreAndResult": 3,
            "winnerOrDrawOnly": 1,
            "miss": 0
        },
        "matches": matches,
        "participants": participants
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT_PATH.relative_to(ROOT)}")
    print(f"Matches: {len(matches)}")
    print(f"Participants: {len(participants)}")


if __name__ == "__main__":
    main()
